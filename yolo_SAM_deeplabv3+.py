import os
from pathlib import Path

import cv2
import numpy as np
import openpyxl as op
import torch
from openpyxl.workbook import Workbook
from tqdm import tqdm

from mmseg.apis import inference_model, init_model
from openpyxl import load_workbook

import test as t
from ultralytics import SAM, YOLO
from ultralytics.data.augment import LetterBox


def load_images(folder):
    images = []
    for filename in os.listdir(folder):
        img_path = os.path.join(folder, filename)
        images.append(img_path)
    return images


def get_or_create_sheet(workbook, sheet_name, headers=None):
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(sheet_name)
        if headers:
            worksheet.append(headers)
    return worksheet


def deeplabv3plus(bath_paths):
    config_file = 'Configs/deeplabv3plus101-d-RepLKNet-nopretrain.py'
    checkpoint_file = 'checkpoint/best_mIoU_iter_148000.pth'

    # 根据配置文件和模型文件建立模型
    model = init_model(config_file, checkpoint_file, device='cuda:0')

    # 在单张图像上测试并可视化
    # img_floder = '/home/ubuntu/pythonProject/mmsegmentation/data/my_dataset/img_dir/test/'
    # img = load_images(img_floder)
    result = inference_model(model, bath_paths)
    # mask = result.pred_sem_seg.data[0].cpu().numpy()
    return result


def yolo_SAM(bath_paths):
    sam = SAM("sam_b.pt")
    yolo_rice = YOLO("runs/segment/train-allrice-ori/weights/best.pt")
    # yolo_chalkiness = YOLO("runs/segment/train_DBB_Revcol_chalkiness_part_all/weights/best.pt")
    # results_chalkiness = yolo_chalkiness.predict(image_folder, conf=0.4)

    # 遍历图片文件夹下的所有图片
    result_SAM = []
    for i, imgPath in enumerate(bath_paths):
        # 读取图片
        img = cv2.imread(imgPath)
        height, width = img.shape[:2]
        img = cv2.resize(img, (int(width *0.3), int(height *0.3)))
        # 获取YOLO检测结果
        yoloR = yolo_rice(img, save=False)[0]
        boxes = yoloR.boxes.xyxy.cpu().numpy()

        if boxes.shape[0] == 0:
            # 没检测到任何框，直接跳过或返回空结果
            result = None  # 或者 []，根据sam函数需求调整
        else:
            result = sam(img, bboxes=boxes)
        result_SAM.append(result)
    return result_SAM




def merge(image, result_SAM, result_deeplabv3plus, imgName):
    rice_color = [255, 0, 0]
    chalkiness_color = [0, 0, 255]
    output_path = f"./yolo_SAM_deeplabv3plus"

    # 调整 SAM 掩码大小
    img = LetterBox(result_SAM.masks.shape[1:])(image=np.asarray(image))
    im_gpu = (
        torch.as_tensor(img, dtype=torch.float16, device=result_SAM.masks.data.device)
        .permute(2, 0, 1)
        .flip(0)
        .contiguous() / 255
    )

    result_SAM_masks = result_SAM.masks.data.int().clone().detach().to(device='cuda')
    result_SAM_masks_ori = t.scale_image(
        torch.sum(result_SAM_masks, dim=0).repeat(1, 1, 1).permute(1, 2, 0).cpu().numpy().astype(np.uint8),
        image.shape
    )   #H×W×1

    result_deeplabv3plus_masks = result_deeplabv3plus.pred_sem_seg.data.clone().detach().to(device='cuda')
    result_deeplabv3plus_masks = torch.from_numpy(result_SAM_masks_ori.transpose((2, 0, 1))).to('cuda') & result_deeplabv3plus_masks

    colors_rice = [rice_color] * len(result_SAM_masks)
    colors_chalkiness = [chalkiness_color] * len(result_deeplabv3plus_masks)

    img = t.masks(result_SAM_masks, colors=colors_rice, im_gpu=im_gpu, im=image)

    img_chalkiness = LetterBox(result_deeplabv3plus.pred_sem_seg.shape[0:])(image=np.asarray(img))
    im_gpu_chalkiness = (
        torch.as_tensor(img_chalkiness, dtype=torch.float16, device=result_SAM.masks.data.device)
        .permute(2, 0, 1)
        .flip(0)
        .contiguous() / 255
    )
    img = t.masks(result_deeplabv3plus_masks, colors=colors_chalkiness, im_gpu=im_gpu_chalkiness, im=img)

    # Excel 表相关处理
    excel_chalkiness = Path(output_path) / 'chalkiness'
    if not os.path.exists(excel_chalkiness):
        os.makedirs(excel_chalkiness)
    excel_path_chalkiness = str(excel_chalkiness / f'{Path(imgName).stem}.xlsx')
    if os.path.exists(excel_path_chalkiness):
        workbook = load_workbook(excel_path_chalkiness)
    else:
        workbook = op.Workbook()
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]
    worksheet1 = get_or_create_sheet(workbook, '垩白度', ['序号', '垩白度'])
    worksheet2 = get_or_create_sheet(workbook, '垩白比例', ['序号', '垩白比例'])

    chalkiness_mask = cv2.resize(result_deeplabv3plus_masks.permute(1, 2, 0).cpu().numpy().astype(np.uint8),
                                 (int(img.shape[1]*0.3), int(img.shape[0]*0.3)))

    chalkiness_ratios = []  # 保存每个米粒的垩白度
    chalkiness_number = 0

    for rice_number in range(len(result_SAM_masks)):
        rice_mask_one = result_SAM_masks[rice_number].cpu().numpy().astype(np.uint8)
        rice_area = rice_mask_one.sum()
        if rice_area == 0:
            worksheet1.append([rice_number, 0])
            chalkiness_ratios.append(0)  # 仍加入 0
            continue

        chalkiness_mask_one = (chalkiness_mask & rice_mask_one)
        chalkiness_area = chalkiness_mask_one.sum()

        ratio = chalkiness_area / rice_area
        chalkiness_ratios.append(ratio)
        worksheet1.append([rice_number, ratio])

        if chalkiness_area > 0:
            chalkiness_number += 1  # 只有有垩白才加 1
        x, y = result_SAM.masks.xy[rice_number][0] * (1 / 0.3)
        cv2.putText(img, f"{rice_number}", (int(x), int(y)), cv2.FONT_HERSHEY_SIMPLEX, 3, (255, 255, 255), 4, cv2.LINE_AA)

    worksheet2.append([1, chalkiness_number / len(result_SAM_masks)])
    workbook.save(excel_path_chalkiness)

    # 计算平均垩白度并更新总汇总表
    average_chalkiness = np.mean(chalkiness_ratios)
    std_chalkiness = np.std(chalkiness_ratios) if chalkiness_ratios else 0
    summary_path = Path(output_path) / "chalkiness_summary.xlsx"
    if summary_path.exists():
        summary_workbook = load_workbook(str(summary_path))
    else:
        summary_workbook = op.Workbook()
        if "Sheet" in summary_workbook.sheetnames:
            del summary_workbook["Sheet"]
    summary_sheet = get_or_create_sheet(summary_workbook, "汇总", ["编号", "图片名", "平均垩白度","垩白度标准差"])
    summary_sheet.append([summary_sheet.max_row, imgName, average_chalkiness,std_chalkiness])
    summary_workbook.save(str(summary_path))

    os.makedirs(output_path, exist_ok=True)
    cv2.imwrite(os.path.join(output_path, f"{imgName}"), img)


if __name__ == "__main__":
    image_folder = "/home/ubuntu/data/chalkiness test/hainanebaigaitest"
    imgPaths = [os.path.join(image_folder, imgName) for imgName in os.listdir(image_folder)]

    # tqdm包装图片路径列表，实现进度条
    for imgPath in tqdm(imgPaths, desc="处理图片"):
        img = cv2.imread(imgPath)

        # 单张推理
        result_SAM = yolo_SAM([imgPath])[0][0]  # 注意：yolo_SAM 返回列表的列表
        result_deeplabv3plus = deeplabv3plus([imgPath])[0]

        merge(img, result_SAM, result_deeplabv3plus, os.path.basename(imgPath))

        # 手动释放缓存，减少内存泄露
        del result_SAM, result_deeplabv3plus, img
        torch.cuda.empty_cache()
